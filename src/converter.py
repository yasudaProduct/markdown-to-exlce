import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any, Optional
import os


class ExcelConverter:
    """MarkdownテーブルデータをExcelファイルに変換するクラス"""
    
    def __init__(self):
        self.default_font = Font(name='Arial', size=10)
        self.header_font = Font(name='Arial', size=10, bold=True)
    
    def convert_to_excel(
        self, 
        tables_data: List[Dict[str, Any]], 
        output_path: str,
        apply_formatting: bool = False,
        auto_adjust_width: bool = False
    ) -> None:
        """
        テーブルデータをExcelファイルに変換する
        
        Args:
            tables_data: テーブルデータのリスト
            output_path: 出力Excelファイルパス
            apply_formatting: フォーマット適用するか
            auto_adjust_width: 列幅自動調整するか
        """
        # 出力ディレクトリの確認
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            raise Exception(f"Output directory does not exist: {output_dir}")
        
        # ワークブック作成
        workbook = openpyxl.Workbook()
        
        # デフォルトシートを削除
        if workbook.active:
            workbook.remove(workbook.active)
        
        if not tables_data:
            # 空のテーブルリストの場合、空のシートを作成
            worksheet = workbook.create_sheet("Sheet1")
        else:
            # 各テーブルに対してシートを作成
            for i, table_data in enumerate(tables_data):
                sheet_name = f"Table{i+1}" if len(tables_data) > 1 else "Sheet1"
                worksheet = workbook.create_sheet(sheet_name)
                
                self._populate_worksheet(
                    worksheet, 
                    table_data, 
                    apply_formatting, 
                    auto_adjust_width
                )
        
        # ファイルに保存
        workbook.save(output_path)
    
    def _populate_worksheet(
        self, 
        worksheet, 
        table_data: Dict[str, Any],
        apply_formatting: bool,
        auto_adjust_width: bool
    ) -> None:
        """
        ワークシートにテーブルデータを設定する
        
        Args:
            worksheet: openpyxlワークシート
            table_data: テーブルデータ
            apply_formatting: フォーマット適用するか
            auto_adjust_width: 列幅自動調整するか
        """
        headers = table_data.get('headers', [])
        rows = table_data.get('rows', [])
        alignment = table_data.get('alignment', [])
        
        # ヘッダーを設定
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            
            if apply_formatting:
                cell.font = self.header_font
                
                # アライメント設定
                if col_idx <= len(alignment):
                    align_type = alignment[col_idx - 1]
                    if align_type == 'center':
                        cell.alignment = Alignment(horizontal='center')
                    elif align_type == 'right':
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='left')
        
        # データ行を設定
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                # 空文字列の場合はNoneに変換
                value = cell_value if cell_value != '' else None
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                if apply_formatting:
                    cell.font = self.default_font
                    
                    # アライメント設定
                    if col_idx <= len(alignment):
                        align_type = alignment[col_idx - 1]
                        if align_type == 'center':
                            cell.alignment = Alignment(horizontal='center')
                        elif align_type == 'right':
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='left')
        
        # 列幅の自動調整
        if auto_adjust_width:
            self._auto_adjust_column_width(worksheet, headers, rows)
    
    def _auto_adjust_column_width(
        self, 
        worksheet, 
        headers: List[str], 
        rows: List[List[str]]
    ) -> None:
        """
        列幅を自動調整する
        
        Args:
            worksheet: openpyxlワークシート
            headers: ヘッダーリスト
            rows: データ行リスト
        """
        for col_idx, header in enumerate(headers, 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            
            # ヘッダーの長さを計算
            max_length = len(str(header))
            
            # 各行のセル値の長さを計算
            for row_data in rows:
                if col_idx <= len(row_data):
                    cell_value = row_data[col_idx - 1]
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
            
            # 最小幅と最大幅を設定
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def convert_from_dataframe(
        self, 
        dataframes: List[pd.DataFrame], 
        output_path: str,
        sheet_names: Optional[List[str]] = None
    ) -> None:
        """
        DataFrameからExcelファイルに変換する（代替メソッド）
        
        Args:
            dataframes: DataFrameのリスト
            output_path: 出力Excelファイルパス
            sheet_names: シート名のリスト
        """
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for i, df in enumerate(dataframes):
                sheet_name = sheet_names[i] if sheet_names and i < len(sheet_names) else f'Sheet{i+1}'
                df.to_excel(writer, sheet_name=sheet_name, index=False)