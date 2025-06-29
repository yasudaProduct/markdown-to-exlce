import pytest
import pandas as pd
import tempfile
import os
from pathlib import Path
from openpyxl import load_workbook
from src.converter import ExcelConverter


class TestExcelConverter:
    
    def test_convert_single_table_to_excel(self):
        """単一テーブルのExcel変換テスト"""
        # テストデータ
        table_data = {
            'headers': ['Name', 'Age', 'City'],
            'rows': [
                ['Alice', '25', 'Tokyo'],
                ['Bob', '30', 'Osaka']
            ],
            'alignment': ['left', 'center', 'left']
        }
        
        converter = ExcelConverter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            try:
                # 変換実行
                converter.convert_to_excel([table_data], tmp_file.name)
                
                # ファイルが作成されたことを確認
                assert os.path.exists(tmp_file.name)
                
                # 内容を検証
                workbook = load_workbook(tmp_file.name)
                sheet = workbook.active
                
                # ヘッダーの確認
                assert sheet['A1'].value == 'Name'
                assert sheet['B1'].value == 'Age'
                assert sheet['C1'].value == 'City'
                
                # データの確認
                assert sheet['A2'].value == 'Alice'
                assert sheet['B2'].value == '25'
                assert sheet['C2'].value == 'Tokyo'
                assert sheet['A3'].value == 'Bob'
                assert sheet['B3'].value == '30'
                assert sheet['C3'].value == 'Osaka'
                
            finally:
                # 一時ファイルを削除
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)
    
    def test_convert_multiple_tables_to_excel(self):
        """複数テーブルのExcel変換テスト（複数シート）"""
        # テストデータ
        tables_data = [
            {
                'headers': ['Product', 'Price'],
                'rows': [['Apple', '100'], ['Orange', '150']],
                'alignment': ['left', 'right']
            },
            {
                'headers': ['Country', 'Capital'],
                'rows': [['Japan', 'Tokyo'], ['USA', 'Washington']],
                'alignment': ['left', 'left']
            }
        ]
        
        converter = ExcelConverter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            try:
                # 変換実行
                converter.convert_to_excel(tables_data, tmp_file.name)
                
                # ファイルが作成されたことを確認
                assert os.path.exists(tmp_file.name)
                
                # 内容を検証
                workbook = load_workbook(tmp_file.name)
                
                # シート数の確認
                assert len(workbook.sheetnames) == 2
                
                # 1つ目のシート確認
                sheet1 = workbook.worksheets[0]
                assert sheet1.title == 'Table1'
                assert sheet1['A1'].value == 'Product'
                assert sheet1['B1'].value == 'Price'
                assert sheet1['A2'].value == 'Apple'
                assert sheet1['B2'].value == '100'
                
                # 2つ目のシート確認
                sheet2 = workbook.worksheets[1]
                assert sheet2.title == 'Table2'
                assert sheet2['A1'].value == 'Country'
                assert sheet2['B1'].value == 'Capital'
                assert sheet2['A2'].value == 'Japan'
                assert sheet2['B2'].value == 'Tokyo'
                
            finally:
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)
    
    def test_convert_empty_table_list(self):
        """空のテーブルリストの処理テスト"""
        converter = ExcelConverter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            try:
                # 空のリストで変換実行
                converter.convert_to_excel([], tmp_file.name)
                
                # ファイルが作成されたことを確認
                assert os.path.exists(tmp_file.name)
                
                # 空のワークブックが作成されることを確認
                workbook = load_workbook(tmp_file.name)
                assert len(workbook.sheetnames) == 1
                sheet = workbook.active
                assert sheet['A1'].value is None
                
            finally:
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)
    
    def test_convert_table_with_empty_cells(self):
        """空セルを含むテーブルの変換テスト"""
        table_data = {
            'headers': ['Name', 'Age', 'City'],
            'rows': [
                ['Alice', '', 'Tokyo'],
                ['', '30', ''],
                ['Carol', '28', 'Kyoto']
            ],
            'alignment': ['left', 'center', 'left']
        }
        
        converter = ExcelConverter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            try:
                converter.convert_to_excel([table_data], tmp_file.name)
                
                workbook = load_workbook(tmp_file.name)
                sheet = workbook.active
                
                # 空セルの確認
                assert sheet['A2'].value == 'Alice'
                assert sheet['B2'].value == '' or sheet['B2'].value is None
                assert sheet['C2'].value == 'Tokyo'
                
                assert sheet['A3'].value == '' or sheet['A3'].value is None
                assert sheet['B3'].value == '30'
                assert sheet['C3'].value == '' or sheet['C3'].value is None
                
            finally:
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)
    
    def test_convert_with_formatting(self):
        """フォーマット設定を含む変換テスト"""
        table_data = {
            'headers': ['Name', 'Score', 'Grade'],
            'rows': [
                ['Alice', '95.5', 'A'],
                ['Bob', '87.2', 'B']
            ],
            'alignment': ['left', 'center', 'right']
        }
        
        converter = ExcelConverter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            try:
                converter.convert_to_excel([table_data], tmp_file.name, apply_formatting=True)
                
                workbook = load_workbook(tmp_file.name)
                sheet = workbook.active
                
                # データの確認
                assert sheet['A1'].value == 'Name'
                assert sheet['B1'].value == 'Score'
                assert sheet['C1'].value == 'Grade'
                
                # フォーマットが適用されていることを確認（ヘッダーが太字など）
                assert sheet['A1'].font.bold == True
                assert sheet['B1'].font.bold == True
                assert sheet['C1'].font.bold == True
                
            finally:
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)
    
    def test_auto_adjust_column_width(self):
        """列幅自動調整テスト"""
        table_data = {
            'headers': ['Very Long Column Name', 'Short', 'Medium Length'],
            'rows': [
                ['This is a very long text content', 'A', 'Medium text'],
                ['Short', 'Another short text', 'Med']
            ],
            'alignment': ['left', 'center', 'right']
        }
        
        converter = ExcelConverter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            try:
                converter.convert_to_excel([table_data], tmp_file.name, auto_adjust_width=True)
                
                workbook = load_workbook(tmp_file.name)
                sheet = workbook.active
                
                # 列幅が調整されていることを確認
                # 正確な幅の値は環境依存なので、調整されていることだけ確認
                assert sheet.column_dimensions['A'].width > 8.43  # デフォルト幅より大きい
                
            finally:
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)
    
    def test_invalid_output_path(self):
        """無効な出力パスのエラーハンドリングテスト"""
        table_data = {
            'headers': ['Name', 'Age'],
            'rows': [['Alice', '25']],
            'alignment': ['left', 'right']
        }
        
        converter = ExcelConverter()
        
        # 存在しないディレクトリへの出力
        invalid_path = '/nonexistent/directory/output.xlsx'
        
        with pytest.raises(Exception):
            converter.convert_to_excel([table_data], invalid_path)