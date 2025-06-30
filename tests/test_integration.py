import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import load_workbook
from src.integration import MarkdownToExcelProcessor


class TestMarkdownToExcelIntegration:
    """Parser + Converter + CLIの統合テスト"""
    
    def test_end_to_end_single_table_conversion(self):
        """単一テーブルのエンドツーエンド変換テスト"""
        markdown_content = """# 販売データ

| 商品名 | 価格 | 在庫数 |
|--------|------|--------|
| りんご | 120 | 50 |
| みかん | 80 | 30 |
| ばなな | 100 | 25 |

上記は今月の販売データです。
"""
        
        processor = MarkdownToExcelProcessor()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            input_file = Path(temp_dir) / "sales.md"
            output_file = Path(temp_dir) / "sales.xlsx"
            
            input_file.write_text(markdown_content, encoding='utf-8')
            
            # エンドツーエンド変換実行
            result = processor.process_file(
                str(input_file),
                str(output_file),
                apply_formatting=True,
                auto_adjust_width=True
            )
            
            # 変換結果の確認
            assert result.success == True
            assert result.tables_found == 1
            assert result.output_file == str(output_file)
            assert output_file.exists()
            
            # Excelファイル内容の検証
            workbook = load_workbook(output_file)
            sheet = workbook.active
            
            # ヘッダーの確認
            assert sheet['A1'].value == '商品名'
            assert sheet['B1'].value == '価格'
            assert sheet['C1'].value == '在庫数'
            
            # データの確認
            assert sheet['A2'].value == 'りんご'
            assert sheet['B2'].value == '120'
            assert sheet['C2'].value == '50'
            
            assert sheet['A3'].value == 'みかん'
            assert sheet['B3'].value == '80'
            assert sheet['C3'].value == '30'
            
            # フォーマットの確認
            assert sheet['A1'].font.bold == True
            assert sheet['B1'].font.bold == True
            assert sheet['C1'].font.bold == True
    
    def test_end_to_end_multiple_tables_conversion(self):
        """複数テーブルのエンドツーエンド変換テスト"""
        markdown_content = """# レポート

## 売上データ

| 月 | 売上 |
|----|------|
| 1月 | 1000 |
| 2月 | 1200 |

## 経費データ

| 項目 | 金額 |
|------|------|
| 家賃 | 500 |
| 光熱費 | 100 |
| 通信費 | 50 |
"""
        
        processor = MarkdownToExcelProcessor()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            input_file = Path(temp_dir) / "report.md"
            output_file = Path(temp_dir) / "report.xlsx"
            
            input_file.write_text(markdown_content, encoding='utf-8')
            
            result = processor.process_file(
                str(input_file),
                str(output_file),
                apply_formatting=False,
                auto_adjust_width=False
            )
            
            # 変換結果の確認
            assert result.success == True
            assert result.tables_found == 2
            assert output_file.exists()
            
            # 複数シートの確認
            workbook = load_workbook(output_file)
            assert len(workbook.sheetnames) == 2
            
            # 1つ目のシート（売上データ）
            sheet1 = workbook.worksheets[0]
            assert sheet1.title == 'Table1'
            assert sheet1['A1'].value == '月'
            assert sheet1['B1'].value == '売上'
            assert sheet1['A2'].value == '1月'
            assert sheet1['B2'].value == '1000'
            
            # 2つ目のシート（経費データ）
            sheet2 = workbook.worksheets[1]
            assert sheet2.title == 'Table2'
            assert sheet2['A1'].value == '項目'
            assert sheet2['B1'].value == '金額'
            assert sheet2['A2'].value == '家賃'
            assert sheet2['B2'].value == '500'
    
    def test_end_to_end_no_tables_handling(self):
        """テーブルなしファイルのエンドツーエンド処理テスト"""
        markdown_content = """# ドキュメント

これはテーブルを含まないドキュメントです。

- リスト項目1
- リスト項目2

## セクション

ただのテキストです。
"""
        
        processor = MarkdownToExcelProcessor()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            input_file = Path(temp_dir) / "no_tables.md"
            output_file = Path(temp_dir) / "no_tables.xlsx"
            
            input_file.write_text(markdown_content, encoding='utf-8')
            
            result = processor.process_file(
                str(input_file),
                str(output_file)
            )
            
            # 変換結果の確認
            assert result.success == True
            assert result.tables_found == 0
            assert result.warnings == ['No tables found in the input file']
            assert output_file.exists()
            
            # 空のExcelファイルが作成されることを確認
            workbook = load_workbook(output_file)
            assert len(workbook.sheetnames) == 1
            sheet = workbook.active
            assert sheet['A1'].value is None
    
    def test_end_to_end_batch_processing(self):
        """バッチ処理のエンドツーエンドテスト"""
        processor = MarkdownToExcelProcessor()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir) / "input"
            output_dir = Path(temp_dir) / "output"
            input_dir.mkdir()
            output_dir.mkdir()
            
            # 複数のテストファイル作成
            (input_dir / "file1.md").write_text("""
| 名前 | 年齢 |
|------|------|
| 太郎 | 25 |
| 花子 | 30 |
""", encoding='utf-8')
            
            (input_dir / "file2.md").write_text("""
| 都市 | 人口 |
|------|------|
| 東京 | 1400 |
| 大阪 | 880 |
""", encoding='utf-8')
            
            (input_dir / "empty.md").write_text("# 空ファイル", encoding='utf-8')
            (input_dir / "not_markdown.txt").write_text("テキストファイル")
            
            # バッチ処理実行
            results = processor.process_directory(
                str(input_dir),
                str(output_dir),
                apply_formatting=True,
                auto_adjust_width=True
            )
            
            # 処理結果の確認
            assert len(results) == 3  # Markdownファイルのみ処理
            assert all(result.success for result in results)
            
            # 出力ファイルの確認
            assert (output_dir / "file1.xlsx").exists()
            assert (output_dir / "file2.xlsx").exists()
            assert (output_dir / "empty.xlsx").exists()
            assert not (output_dir / "not_markdown.xlsx").exists()
            
            # file1.xlsxの内容確認
            workbook1 = load_workbook(output_dir / "file1.xlsx")
            sheet1 = workbook1.active
            assert sheet1['A1'].value == '名前'
            assert sheet1['A2'].value == '太郎'
    
    def test_end_to_end_malformed_table_handling(self):
        """不正形式テーブルのエンドツーエンド処理テスト"""
        markdown_content = """# 不正テーブルテスト

| 名前 | 年齢 | 都市 |
|------|------|------|
| 太郎 | 25 | 東京 | 余分 |
| 花子 |
| 次郎 | 30 |

正常なテーブル:

| A | B |
|---|---|
| 1 | 2 |
"""
        
        processor = MarkdownToExcelProcessor()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            input_file = Path(temp_dir) / "malformed.md"
            output_file = Path(temp_dir) / "malformed.xlsx"
            
            input_file.write_text(markdown_content, encoding='utf-8')
            
            result = processor.process_file(
                str(input_file),
                str(output_file)
            )
            
            # 不正形式でも処理が成功することを確認
            assert result.success == True
            assert result.tables_found == 2
            assert output_file.exists()
            
            # データが適切に正規化されていることを確認
            workbook = load_workbook(output_file)
            
            # 1つ目のシート（不正形式テーブル）
            sheet1 = workbook.worksheets[0]
            assert sheet1['A1'].value == '名前'
            assert sheet1['B1'].value == '年齢'
            assert sheet1['C1'].value == '都市'
            
            # 不正行も適切に処理されている
            assert sheet1['A2'].value == '太郎'
            assert sheet1['B2'].value == '25'
            assert sheet1['C2'].value == '東京'  # 余分な列は切り捨て
            
            assert sheet1['A3'].value == '花子'
            assert sheet1['B3'].value == '' or sheet1['B3'].value is None
            assert sheet1['C3'].value == '' or sheet1['C3'].value is None
    
    def test_end_to_end_performance_large_table(self):
        """大きなテーブルのパフォーマンステスト"""
        # 1000行のテーブルを作成
        headers = "| ID | 名前 | 値 |\n|----|----|----|\n"
        rows = []
        for i in range(1000):
            rows.append(f"| {i} | Item{i} | {i * 10} |")
        
        markdown_content = "# 大きなテーブル\n\n" + headers + "\n".join(rows)
        
        processor = MarkdownToExcelProcessor()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            input_file = Path(temp_dir) / "large.md"
            output_file = Path(temp_dir) / "large.xlsx"
            
            input_file.write_text(markdown_content, encoding='utf-8')
            
            import time
            start_time = time.time()
            
            result = processor.process_file(
                str(input_file),
                str(output_file)
            )
            
            end_time = time.time()
            processing_time = end_time - start_time
            
            # パフォーマンスとデータの確認
            assert result.success == True
            assert result.tables_found == 1
            assert processing_time < 10.0  # 10秒以内で完了
            assert output_file.exists()
            
            # データサイズの確認
            workbook = load_workbook(output_file)
            sheet = workbook.active
            assert sheet['A1'].value == 'ID'
            assert sheet['A1001'].value == '999'  # 最後の行
            assert sheet['C1001'].value == '9990'
    
    def test_end_to_end_unicode_handling(self):
        """Unicode文字のエンドツーエンド処理テスト"""
        markdown_content = """# 国際化テスト

| 言語 | 挨拶 | 絵文字 |
|------|------|--------|
| 日本語 | こんにちは | 🗾 |
| English | Hello | 🇺🇸 |
| Français | Bonjour | 🇫🇷 |
| 中文 | 你好 | 🇨🇳 |
| العربية | مرحبا | 🇸🇦 |
| Русский | Привет | 🇷🇺 |
"""
        
        processor = MarkdownToExcelProcessor()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            input_file = Path(temp_dir) / "unicode.md"
            output_file = Path(temp_dir) / "unicode.xlsx"
            
            input_file.write_text(markdown_content, encoding='utf-8')
            
            result = processor.process_file(
                str(input_file),
                str(output_file)
            )
            
            # Unicode処理の確認
            assert result.success == True
            assert result.tables_found == 1
            assert output_file.exists()
            
            # Unicode文字が正しく保存されていることを確認
            workbook = load_workbook(output_file)
            sheet = workbook.active
            
            assert sheet['A1'].value == '言語'
            assert sheet['B1'].value == '挨拶'
            assert sheet['C1'].value == '絵文字'
            
            assert sheet['A2'].value == '日本語'
            assert sheet['B2'].value == 'こんにちは'
            assert sheet['C2'].value == '🗾'
            
            assert sheet['A6'].value == 'العربية'
            assert sheet['B6'].value == 'مرحبا'
            assert sheet['C6'].value == '🇸🇦'